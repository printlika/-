import requests
import openpyxl
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import threading

# 配置登录的 URL 和表单数据
login_url = ""  # 替换为实际登录URL

def start_login():
    try:
        # 获取用户名和密码列表
        users = []
        passwords = []

        # 打开文本文件
        with open("", "r") as file:
            lines = file.readlines()

        # 创建一个新的Excel工作簿
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # 将文本文件的内容写入Excel工作表
        for row, line in enumerate(lines, start=1):
            for col, value in enumerate(line.split(), start=1):
                worksheet.cell(row=row, column=col, value=value)

        # 保存Excel文件
        workbook.save(".xlsx")

        # 打开Excel文件并选择默认的工作表
        workbook = openpyxl.load_workbook(".xlsx")
        worksheet = workbook.active

        # 遍历工作表中的每一行
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            username, password = row
            users.append(username)
            passwords.append(password)

        # 打印用户名和密码列表
        print("用户名列表:", users)
        print("密码列表:", passwords)

        # 初始化进度条
        progress['value'] = 0
        progress['maximum'] = len(users)

        # 创建一个新的Excel工作簿来保存结果
        result_workbook = openpyxl.Workbook()
        result_worksheet = result_workbook.active
        result_worksheet.append(["用户名", "密码", "状态"])

        # 登录并记录结果
        for index, (username, password) in enumerate(zip(users, passwords), start=1):
            # 发送POST请求进行登录
            response = requests.post(login_url, data={"username": username, "password": password})

            # 检查返回的状态码
            if response.status_code == 200:
                status = "登录成功"
            else:
                status = "登录失败"

            # 写入一行数据
            result_worksheet.append([username, password, status])

            # 更新进度条
            progress['value'] = index

        # 保存结果Excel文件
        result_workbook.save("结果.xlsx")
        print("任务完成！")

        # 删除重复内容
        workbook = openpyxl.load_workbook("结果.xlsx")
        worksheet = workbook.active
        for row in range(worksheet.max_row, 1, -1):
            for col in range(1, worksheet.max_column + 1):
                if worksheet.cell(row=row, column=col).value == worksheet.cell(row=row-1, column=col).value:
                    worksheet.delete_rows(row)
                    break
        workbook.save("结果.xlsx")
        print("重复内容删除完成！")

    except Exception as e:
        messagebox.showerror("错误", str(e))

def run_login_in_thread():
    # 在新线程中运行登录函数，以避免阻塞主线程
    threading.Thread(target=start_login).start()

def select_dictionary():
    # 打开文件对话框选择字典文件
    file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
    if file_path:
        # 更新字典文件路径
        dictionary_path.set(file_path)

def export_results():
    # 打开文件对话框选择导出结果文件
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if file_path:
        # 导出结果到文件
        # 这里需要添加导出结果的代码
        messagebox.showinfo("导出", "结果已导出到 " + file_path)

def import_dictionary():
    # 打开文件对话框选择导入字典文件
    file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
    if file_path:
        # 导入字典文件
        # 这里需要添加导入字典文件的代码
        messagebox.showinfo("导入", "字典已导入 " + file_path)

def export_dictionary():
    # 打开文件对话框选择导出字典文件
    file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
    if file_path:
        # 导出字典文件
        # 这里需要添加导出字典文件的代码
        messagebox.showinfo("导出", "字典已导出到 " + file_path)

# 创建主窗口
root = tk.Tk()
root.title("登录工具")

# 创建一个标签和按钮来选择字典文件
dictionary_path = tk.StringVar()
dictionary_label = tk.Label(root, text="字典文件路径:")
dictionary_label.pack()
dictionary_entry = tk.Entry(root, textvariable=dictionary_path, width=50)
dictionary_entry.pack()
dictionary_button = tk.Button(root, text="选择字典", command=select_dictionary)
dictionary_button.pack()

# 创建一个按钮来启动登录过程
start_button = tk.Button(root, text="开始登录", command=run_login_in_thread)
start_button.pack(pady=20)

# 创建一个进度条
progress = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
progress.pack(pady=20)

# 创建一个按钮来导出结果
export_button = tk.Button(root, text="导出结果", command=export_results)
export_button.pack(pady=20)

# 创建一个按钮来导入字典文件
import_button = tk.Button(root, text="导入字典", command=import_dictionary)
import_button.pack(pady=20)

# 创建一个按钮来导出字典文件
export_dict_button = tk.Button(root, text="导出字典", command=export_dictionary)
export_dict_button.pack(pady=20)

# 运行主循环
root.mainloop()
